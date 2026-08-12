import os
import re
import sys
import json
import base64
from typing import Dict, Any
from dotenv import load_dotenv
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables automatically from .env
load_dotenv()

EXPECTED_KEYS = [
    "claim_sui_account_number",
    "claimant_ssn",
    "claimant_name",
    "claim_start_date",
    "claim_end_date",
    "byb_date",
    "bye_date",
    "claim_mailing_date",
    "claim_liability_percentage",
    "claim_liability_base_amount",
    "agency_address_line_1",
    "agency_address_line_2",
    "separation_code"
]


# ==========================================
# Data Normalization Helpers for Validation
# ==========================================

def normalize_ssn(val: str) -> str:
    """Strips non-digit characters to compare 9-digit SSNs."""
    if not val:
        return ""
    digits = re.sub(r"\D", "", str(val))
    return digits if len(digits) == 9 else str(val).strip()


def normalize_date(val: str) -> str:
    """Converts MM/DD/YY or MM/DD/YYYY to standard MM/DD/YYYY."""
    if not val:
        return ""
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", str(val))
    if match:
        m, d, y = match.groups()
        m = m.zfill(2)
        d = d.zfill(2)
        if len(y) == 2:
            y = f"20{y}"
        return f"{m}/{d}/{y}"
    return str(val).strip()


def normalize_amount(val: str) -> str:
    """Formats number strings to float representation for clean comparison."""
    if not val:
        return ""
    cleaned = re.sub(r"[^\d.]", "", str(val))
    if not cleaned:
        return str(val).strip()
    try:
        num = float(cleaned)
        return f"{num:.2f}"
    except ValueError:
        return str(val).strip()


def normalize_sui_account(val: str) -> str:
    """Pads SUI Account Number with leading zeros to standard 10 digits."""
    if not val:
        return ""
    digits = re.sub(r"\D", "", str(val))
    if digits:
        return digits.zfill(10)
    return str(val).strip()


def normalize_general(val: str) -> str:
    """General text normalization: uppercase and single spacing."""
    if not val:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip().upper()


def validate_field(pdf_val: str, ocr_val: str, field_name: str) -> str:
    """Compares normalized values of PDF vs OCR for a field."""
    if not pdf_val and not ocr_val:
        return "Match"
    if not pdf_val or not ocr_val:
        return "Mismatch"

    norm_pdf = pdf_val
    norm_ocr = ocr_val

    if "ssn" in field_name:
        norm_pdf = normalize_ssn(pdf_val)
        norm_ocr = normalize_ssn(ocr_val)
    elif "date" in field_name:
        norm_pdf = normalize_date(pdf_val)
        norm_ocr = normalize_date(ocr_val)
    elif "amount" in field_name:
        norm_pdf = normalize_amount(pdf_val)
        norm_ocr = normalize_amount(ocr_val)
    elif "account" in field_name:
        norm_pdf = normalize_sui_account(pdf_val)
        norm_ocr = normalize_sui_account(ocr_val)
    elif "percentage" in field_name:
        norm_pdf = normalize_amount(pdf_val)
        norm_ocr = normalize_amount(ocr_val)
    elif "separation" in field_name:
        norm_pdf = normalize_general(pdf_val)
        norm_ocr = normalize_general(ocr_val)
        code_pdf = re.split(r"[\s\-]", norm_pdf)[0] if norm_pdf else ""
        code_ocr = re.split(r"[\s\-]", norm_ocr)[0] if norm_ocr else ""
        if code_pdf and code_ocr and code_pdf == code_ocr:
            return "Match"
    else:
        norm_pdf = normalize_general(pdf_val)
        norm_ocr = normalize_general(ocr_val)

    return "Match" if norm_pdf == norm_ocr else "Mismatch"


# ==========================================
# OpenAI Vision Dual Extraction Engine
# ==========================================

import io
from PIL import Image

def image_to_base64_crop(pil_img: Image.Image) -> str:
    """Converts a PIL Image object to base64 string."""
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def extract_dual_json_from_image(image_path: str, api_key: str = None, model: str = None) -> Dict[str, Any]:
    """
    Sends the screenshot image to GPT Vision to extract two JSON payloads:
    1. 'pdf': Extracted from high-resolution crop of the left side document notice.
    2. 'ocr': Extracted from high-resolution crop of the right side Entry Screen UI form.
    """
    selected_model = model or os.environ.get("OPENAI_MODEL", "gpt-5.6")
    client = OpenAI(api_key=api_key or os.environ.get("OPENAI_API_KEY"))
    if not client.api_key:
        raise ValueError("OpenAI API Key not found. Please set OPENAI_API_KEY in .env file.")

    # Open image and crop left (PDF) and right (UI Entry Screen) panels for maximum resolution
    im = Image.open(image_path)
    w, h = im.size

    pdf_crop = im.crop((0, 0, int(w * 0.52), h))
    ocr_crop = im.crop((int(w * 0.48), 0, w, h))

    pdf_b64 = image_to_base64_crop(pdf_crop)
    ocr_b64 = image_to_base64_crop(ocr_crop)

    system_prompt = (
        "You are an ultra-high precision AI data extraction assistant specialized in analyzing dual-panel document screenshots.\n"
        "You are provided with TWO cropped images from a split-screen:\n"
        " - IMAGE 1 (PDF / Source Document Notice): Contains the original unemployment notice printout on the LEFT side.\n"
        " - IMAGE 2 (OCR / Entry Screen UI): Contains form fields with entered data on the RIGHT side.\n\n"
        "Extract the following 13 fields from BOTH panels into a single valid JSON object with top-level keys 'pdf' and 'ocr':\n"
        "Required Fields:\n"
        " - claim_sui_account_number\n"
        " - claimant_ssn\n"
        " - claimant_name\n"
        " - claim_start_date\n"
        " - claim_end_date\n"
        " - byb_date\n"
        " - bye_date\n"
        " - claim_mailing_date\n"
        " - claim_liability_percentage (Look under PCT header in Image 1 PDF, e.g. 1.608 or 8.852; look in % of $ box in Image 2 OCR UI, e.g. 1.61 or 8.85)\n"
        " - claim_liability_base_amount\n"
        " - agency_address_line_1 (Look at the Agency / Local Office address section at the bottom of the notice, e.g. 830 PUNCHBOWL #324)\n"
        " - agency_address_line_2 (e.g. HONOLULU, HI 96813-5080)\n"
        " - separation_code (e.g. B or 6 - DISCHARGED NO MISCONDUCT)\n\n"
        "CRITICAL OCR ACCURACY RULES:\n"
        "1. Inspect low-resolution dot-matrix text and cyan/blue-highlighted boxes in Image 1 with extreme pixel care.\n"
        "2. Pay close attention to distinguishing digits like '8' vs '5' vs '6' vs '0' in cyan-highlighted boxes (e.g., under PCT header, '1.608' must be transcribed as 1.608, NOT 1.500; '8.852' must be transcribed as 8.852, NOT 8.562).\n"
        "3. Under key 'pdf', extract exact verbatim values from IMAGE 1.\n"
        "4. Under key 'ocr', extract exact verbatim values from IMAGE 2.\n"
        "5. Output ONLY raw JSON matching this schema without markdown block formatting or chat intro."
    )

    request_params = {
        "model": selected_model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Extract 'pdf' data from Image 1 (PDF Notice) and 'ocr' data from Image 2 (UI Entry Screen) as JSON:"
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{pdf_b64}",
                            "detail": "high"
                        }
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{ocr_b64}",
                            "detail": "high"
                        }
                    }
                ]
            }
        ],
        "response_format": {"type": "json_object"}
    }

    # Only include temperature=0.0 for models that support custom temperature values
    if not (selected_model.startswith("gpt-5") or selected_model.startswith("o1") or selected_model.startswith("o3")):
        request_params["temperature"] = 0.0

    try:
        response = client.chat.completions.create(**request_params)
    except Exception as e:
        # Fallback if API rejects temperature parameter dynamically
        if "temperature" in request_params and "temperature" in str(e).lower():
            del request_params["temperature"]
            response = client.chat.completions.create(**request_params)
        else:
            raise e

    content = response.choices[0].message.content or "{}"
    raw_data = json.loads(content)

    # Standardize data dictionary
    pdf_extracted = {k: raw_data.get("pdf", {}).get(k, "") for k in EXPECTED_KEYS}
    ocr_extracted = {k: raw_data.get("ocr", {}).get(k, "") for k in EXPECTED_KEYS}

    # Perform field-by-field validation
    validation_results = {}
    for key in EXPECTED_KEYS:
        validation_results[key] = validate_field(pdf_extracted[key], ocr_extracted[key], key)

    # Additional feature: Calculate Claim Liability based on OCR values (PCT & MBA)
    calc_liability = compute_calculated_claim_liability(ocr_extracted)
    if calc_liability:
        ocr_extracted["calculated_claim_liability"] = calc_liability
        validation_results["calculated_claim_liability"] = "Calculated"

    # Track token usage and compute per-case cost
    case_basename = os.path.splitext(os.path.basename(image_path))[0]
    token_usage_info = {}
    try:
        from token_monitor import track_usage
        usage_obj = getattr(response, "usage", None)
        token_usage_info = track_usage(usage_obj, model=selected_model, case_name=case_basename)
    except Exception as err:
        print(f"Warning: Could not track token usage: {err}")

    return {
        "pdf": pdf_extracted,
        "ocr": ocr_extracted,
        "validation": validation_results,
        "token_usage": token_usage_info
    }


def compute_calculated_claim_liability(ocr_data: Dict[str, str]) -> str:
    """
    Calculates Claim Liability based on OCR values:
    If OCR PCT == 100% -> $19,890.00
    If OCR PCT < 100%  -> (OCR MBA * OCR PCT) / 100
    """
    pct_val = ocr_data.get("claim_liability_percentage", "")
    mba_val = ocr_data.get("claim_liability_base_amount", "")

    if not pct_val or not mba_val:
        return ""

    try:
        pct_clean = re.sub(r"[^\d.]", "", str(pct_val))
        mba_clean = re.sub(r"[^\d.]", "", str(mba_val))

        if not pct_clean or not mba_clean:
            return ""

        pct = float(pct_clean)
        mba = float(mba_clean)

        if pct >= 100.0:
            return "19890.00"
        else:
            calc_val = (mba * pct) / 100.0
            return f"{calc_val:.2f}"
    except Exception:
        return ""


# ==========================================
# Excel Generator (5-Column Format)
# ==========================================

def export_to_excel(extracted_data: Dict[str, Any], output_excel_path: str):
    """
    Creates an Excel spreadsheet (.xlsx) matching the requested layout:
    Col A: PDF | Col B: PDF Value | Col C: OCR | Col D: OCR Value | Col E: Validation
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"
    ws.views.sheetView[0].showGridLines = True

    # Styling setup
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    normal_font = Font(name="Segoe UI", size=10)
    
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Light green
    mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light red
    match_font = Font(name="Segoe UI", size=10, bold=True, color="375623")
    mismatch_font = Font(name="Segoe UI", size=10, bold=True, color="C65911")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Write Header Row
    headers = ["PDF", "PDF Value", "OCR", "OCR Value", "Validation"]
    ws.append(headers)

    for col_num, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # 2. Write Data Rows
    pdf_data = extracted_data.get("pdf", {})
    ocr_data = extracted_data.get("ocr", {})
    val_data = extracted_data.get("validation", {})

    for idx, key in enumerate(EXPECTED_KEYS, start=2):
        pdf_val = pdf_data.get(key, "")
        ocr_val = ocr_data.get(key, "")
        status = val_data.get(key, "Mismatch")

        ws.cell(row=idx, column=1, value=key).font = bold_font
        ws.cell(row=idx, column=2, value=pdf_val).font = normal_font
        ws.cell(row=idx, column=3, value=key).font = bold_font
        ws.cell(row=idx, column=4, value=ocr_val).font = normal_font

        val_cell = ws.cell(row=idx, column=5, value=f"{status}")
        if status == "Match":
            val_cell.fill = match_fill
            val_cell.font = match_font
        else:
            val_cell.fill = mismatch_fill
            val_cell.font = mismatch_font

        for col_i in range(1, 6):
            c = ws.cell(row=idx, column=col_i)
            c.border = thin_border
            c.alignment = Alignment(vertical="center")

    # Write additional calculated_claim_liability row if present
    if "calculated_claim_liability" in ocr_data:
        calc_row = len(EXPECTED_KEYS) + 2
        calc_key = "calculated_claim_liability"
        calc_val = ocr_data.get(calc_key, "")

        ws.cell(row=calc_row, column=1, value=calc_key).font = bold_font
        ws.cell(row=calc_row, column=2, value="").font = normal_font
        ws.cell(row=calc_row, column=3, value=calc_key).font = bold_font
        ws.cell(row=calc_row, column=4, value=calc_val).font = normal_font

        val_cell = ws.cell(row=calc_row, column=5, value="Calculated")
        val_cell.fill = match_fill
        val_cell.font = match_font

        for col_i in range(1, 6):
            c = ws.cell(row=calc_row, column=col_i)
            c.border = thin_border
            c.alignment = Alignment(vertical="center")

    # 3. Auto-adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    os.makedirs(os.path.dirname(os.path.abspath(output_excel_path)), exist_ok=True)
    try:
        wb.save(output_excel_path)
    except PermissionError:
        alt_path = output_excel_path.replace(".xlsx", "_updated.xlsx")
        print(f"Warning: '{output_excel_path}' is open in another program (Excel). Saving to '{alt_path}' instead.")
        wb.save(alt_path)


# ==========================================
# Core Processing Controller
# ==========================================

def process_screenshot_file(image_path: str, output_dir: str = None) -> Dict[str, Any]:
    """Processes a single screenshot image file."""
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")

    filename = os.path.basename(image_path)
    basename, _ = os.path.splitext(filename)

    base_dir = os.path.dirname(os.path.abspath(image_path))
    target_out_dir = output_dir or os.path.join(base_dir, "output", basename)
    os.makedirs(target_out_dir, exist_ok=True)

    json_output_path = os.path.join(target_out_dir, f"{basename}.json")
    excel_output_path = os.path.join(target_out_dir, f"{basename}.xlsx")

    print(f"\n[Processing] Image: '{filename}'")
    print(f"Extracting 'pdf' and 'ocr' JSON fields via OpenAI Vision...")

    result_data = extract_dual_json_from_image(image_path)

    # Save JSON File (single file containing pdf and ocr JSONs + validation)
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(result_data, f, indent=2)

    # Save Excel File
    export_to_excel(result_data, excel_output_path)

    print(f"Extraction complete for '{filename}'!")
    print(f"Outputs generated in: {target_out_dir}")
    print(f"  - JSON File : {json_output_path}")
    print(f"  - Excel File: {excel_output_path}\n")

    return result_data


if __name__ == "__main__":
    if len(sys.argv) > 1:
        target = sys.argv[1]
        out_directory = sys.argv[2] if len(sys.argv) > 2 else None

        if os.path.isdir(target):
            print(f"Batch processing images in directory: {target}")
            image_extensions = (".png", ".jpg", ".jpeg")
            for fname in os.listdir(target):
                if fname.lower().endswith(image_extensions):
                    img_file = os.path.join(target, fname)
                    process_screenshot_file(img_file, out_directory)
        else:
            process_screenshot_file(target, out_directory)
    else:
        # Default fallback test on Unemployment Claims screenshot
        default_sample = os.path.join(
            "Unemployment Claims - Automation Project",
            "CHOCK II, COLIN K - Entry Screen.png"
        )
        if os.path.exists(default_sample):
            process_screenshot_file(default_sample)
        else:
            print("Usage: python main.py <path_to_image_or_folder> [output_directory]")
